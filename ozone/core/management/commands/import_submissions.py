"""Import Submission from Excel file.
"""
import pickle
import decimal
import logging
import collections

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.timezone import make_aware

from openpyxl import load_workbook

from ozone.core.models import (
    User,
    Party,
    Substance,
    Submission,
    Article7Import,
    Article7Export,
    ReportingPeriod,
    Article7Production,
    Article7Destruction,
    Article7NonPartyTrade,
    Article7Questionnaire,
    ReportingChannel,
    SubmissionFormat,
    Blend,
)

logger = logging.getLogger(__name__)
CACHE_LOC = "/var/tmp/legacy_submission.cache"


class Command(BaseCommand):
    help = "Import Submission from Excel file"
    # Used for double checks.
    import_types = (
        "ImpNew",
        "ImpRecov",
        "ImpFeedstock",
        "ImpEssenUse",
        "ImpProcAgent",
        "ImpQuarAppl",
        "ImpLabUse",
        "ImpPolyol",
    )

    nonparty_types = (
        "NPTImp",
        "NPTExp",
    )

    data_to_check = (
        "imports",
        "exports",
        "produced",
        "destroyed",
        "nonparty",
    )

    def __init__(self, stdout=None, stderr=None, no_color=False):
        super().__init__(stdout=None, stderr=None, no_color=False)

        # Load all values in memory for faster lookups.
        self.current_submission = set(Submission.objects.filter(obligation_id=1).values_list(
            "party__abbr", "reporting_period__name"
        ))
        self.periods = {_period.name: _period
                        for _period in ReportingPeriod.objects.all()}
        self.parties = {_party.abbr: _party
                        for _party in Party.objects.all()}
        self.substances = {_substance.substance_id: _substance
                           for _substance in Substance.objects.all()}
        self.blends = {_blend.legacy_blend_id: _blend
                       for _blend in Blend.objects.all()}

        self.precision = 10

    def add_arguments(self, parser):
        parser.add_argument('file',
                            help="the xlsx input file")
        parser.add_argument('--recreate', action="store_true", default=False,
                            help="Re-create if submission already exists.")
        parser.add_argument('--purge', action="store_true", default=False,
                            help="Purge all entries that were imported")
        parser.add_argument('-l', '--limit', type=int, default=None,
                            help="Limit the number of row to import.")
        parser.add_argument('-p', '--precision', type=int, default=10,
                            help="Number of digits after after the period to "
                                 "check for consistency.")
        parser.add_argument('-C', '--use-cache', action="store_true", default=False,
                            help="Load the data from the cache (if available) instead "
                                 "of the xls")
        parser.add_argument("--dry-run", action="store_true", default=False,
                            help="Only parse the data, but do not insert it.")
        parser.add_argument("-S", "--single", help="Only process this single entry.")

    def process_entry(self, party, period, values, recreate=False, purge=False):
        """Process the parsed data and insert it into the DB.

        Only a wrapper, see _process_entry.
        """
        try:
            return self._process_entry(party, period, values, recreate=recreate, purge=purge)
        except Exception as e:
            logger.error("Error %s while saving: %s/%s", e, party.abbr, period.name,
                         exc_info=True)
            return False

    def get_imports(self, row, party, period):
        """Parses the import data for the submission identified by
        the party/period combination.

        The same data data is duplicated in legacy Excel file. One
        version has the source_party information, and the other one
        only has the total for each submission. Because the source
        party information was introduced at a later time, and made
        optional.

        Double check that the data matches in both sheets and log
        warning if not.

        Return a list of imports.
        """
        imports = []

        # Double check the data from old import sheet with the
        # data from the new import sheet.
        double_check_new = collections.defaultdict(decimal.Decimal)
        double_check_old = collections.defaultdict(decimal.Decimal)

        for import_row in row["ImportNew"]:

            for imp_type in self.import_types:
                pk = party.abbr, period.name, import_row["SubstID"], imp_type
                double_check_new[pk] += decimal.Decimal(import_row[imp_type] or 0)

            if not any(import_row[_npt_type]
                       for _npt_type in ("ImpNew", "ImpRecov")):
                logger.warning("ImportNew no quantity specified: %s/%s/%s", party.abbr,
                               period.name, import_row["SubstID"])

            # Get the source party, if not present then add it as NULL
            # the data on the source party will be in the remarks.
            source_party = import_row["OrgCntryID"].upper()
            if source_party in ("ZZB", "UNK"):
                source_party_id = None
            else:
                try:
                    source_party_id = self.parties[source_party].id
                except KeyError as e:
                    logger.error("Import new unknown source party %s: %s/%s", e, party.abbr,
                                 period.name)
                    source_party_id = None

            substance = self.substances.get(import_row["SubstID"])
            substance_id = None
            try:
                substance_id = substance.id
            except AttributeError:
                pass

            blend = None
            blend_id = None
            if not substance:
                blend = self.blends.get(import_row["SubstID"])
                try:
                    blend_id = blend.id
                except AttributeError:
                    logger.error("Import new unknown substance %s/%s", party.abbr, period.name)
                    continue

            if substance and substance.substance_id != 194:
                quantity_essen_uses = import_row["ImpEssenUse"]
                quantity_crit_uses = None
            else:
                quantity_essen_uses = None
                quantity_crit_uses = import_row["ImpEssenUse"]

            quantity_lab_uses = import_row["ImpLabUse"]
            decision_lab_uses = ""
            decision_essen_uses = ""
            decision_crit_uses = ""
            remark = import_row["Remark"] if import_row["Remark"] is not None else ''
            if quantity_lab_uses and import_row["ImpEssenUse"]:
                diff = import_row["ImpEssenUse"] - quantity_lab_uses
                if diff > 0.0001:
                    decision_lab_uses = remark
                    if substance and substance.substance_id != 194:
                        quantity_essen_uses = diff
                        decision_essen_uses = remark
                    else:
                        quantity_crit_uses = diff
                        decision_crit_uses = remark
                elif -0.0001 <= diff <= 0.0001:
                    quantity_lab_uses = quantity_essen_uses if diff >= 0 else quantity_lab_uses
                    quantity_essen_uses = None
                    decision_lab_uses = remark
                else:
                    logger.warning(
                        "Import: ImpLabUse is greater than ImpEssenUse for %s/%s/%s",
                        party.abbr, period.name, substance.substance_id if substance else 'Unknown substance'
                    )

            imports.append({
                "remarks_party": import_row["Remark"] or "",
                # "remarks_os": "",
                "source_party_id": source_party_id,
                "quantity_total_new": import_row["ImpNew"],
                "quantity_total_recovered": import_row["ImpRecov"],
                "quantity_feedstock": import_row["ImpFeedstock"],
                "quantity_critical_uses": quantity_crit_uses,
                "quantity_essential_uses": quantity_essen_uses,
                "quantity_high_ambient_temperature": None,
                "quantity_laboratory_analytical_uses": quantity_lab_uses,
                "quantity_process_agent_uses": import_row["ImpProcAgent"],
                "quantity_quarantine_pre_shipment": import_row["ImpQuarAppl"],
                "quantity_polyols": import_row["ImpPolyol"],
                "quantity_other_uses": None,
                "decision_critical_uses": decision_crit_uses,
                "decision_essential_uses": decision_essen_uses,
                "decision_high_ambient_temperature": "",
                "decision_laboratory_analytical_uses": decision_lab_uses,
                "decision_process_agent_uses": "",
                "decision_quarantine_pre_shipment": "",
                "decision_other_uses": "",
                "blend_id": blend_id if blend else None,
                # "blend_item_id": "", # created automatically at save()
                "substance_id": substance_id if substance else None,
                # "ordering_id": "",
                # "submission_id": "", # Automatically filled.
            })

        # Cross-Reference with the other sheet, for a double check of data
        # consistency. Only use the ImportNew sheet for now.
        for import_row in row["Import"]:
            for imp_type in self.import_types:
                pk = party.abbr, period.name, import_row["SubstID"], imp_type
                double_check_old[pk] += decimal.Decimal(import_row[imp_type] or 0)

        self.double_check(double_check_old, double_check_new, "Import")

        return imports

    def get_nonparty(self, row, party, period):
        """Parses the "nonparty" data for the submission identified by
        the party/period combination.

        The same data data is duplicated in legacy Excel file. One
        version has the source_party information, and the other one
        only has the total for each submission. Because the source
        party information was introduced at a later time, and made
        optional.

        Double check that the data matches in both sheets and log
        warning if not.

        Returns a list of nonparty trade substances.
        """
        nonparty = []

        # Double check the data from old import sheet with the
        # data from the new import sheet.
        double_check_new = collections.defaultdict(decimal.Decimal)
        double_check_old = collections.defaultdict(decimal.Decimal)

        nonparty_remarks_secretariat = []
        for nonparty_row in row["NonPartyTradeNew"]:
            pk = party.abbr, period.name, nonparty_row["SubstID"], "NPTImp"
            double_check_new[pk] += decimal.Decimal(nonparty_row["NPTImpNew"] or 0)
            double_check_new[pk] += decimal.Decimal(nonparty_row["NPTImpRecov"] or 0)
            pk = party.abbr, period.name, nonparty_row["SubstID"], "NPTExp"
            double_check_new[pk] += decimal.Decimal(nonparty_row["NPTExpNew"] or 0)
            double_check_new[pk] += decimal.Decimal(nonparty_row["NPTExpRecov"] or 0)

            try:
                substance_id = self.substances[nonparty_row["SubstID"]].id
            except KeyError as e:
                logger.error("NonPartyTrade unknown substance %s: %s/%s", e, party.abbr, period.name)
                continue

            if not any(nonparty_row[_npt_type]
                       for _npt_type in ("NPTImpNew", "NPTImpRecov", "NPTExpNew", "NPTExpRecov")):
                if nonparty_row['Remark']:
                    remark = self.substances[nonparty_row["SubstID"]].name + ': ' + nonparty_row['Remark']
                    if remark not in nonparty_remarks_secretariat:
                        nonparty_remarks_secretariat.append(remark)
                logger.warning("NonPartyTradeNew no quantity specified: %s/%s/%s", party.abbr,
                               period.name, nonparty_row["SubstID"])

            # Get the trade party, if not present then add it as NULL
            # the data on the trade party will be in the remarks.
            trade_party = nonparty_row["SrcDestCntryID"].upper()
            if trade_party in ("ZZB", "UNK"):
                trade_party_id = None
            else:
                try:
                    trade_party_id = self.parties[trade_party].id
                except KeyError as e:
                    logger.error("NonPartyTrade new unknown trade party %s: %s/%s", e, party.abbr,
                                 period.name)
                    trade_party_id = None

            nonparty.append({
                "remarks_party": nonparty_row["Remark"] or "",
                "substance_id": substance_id,
                "trade_party_id": trade_party_id,
                # "remarks_os": "",
                "quantity_import_new": nonparty_row["NPTImpNew"],
                "quantity_import_recovered": nonparty_row["NPTImpRecov"],
                "quantity_export_new": nonparty_row["NPTExpNew"],
                "quantity_export_recovered": nonparty_row["NPTExpRecov"],
                # "blend_id": "",
                # "blend_item_id": "",
                # "submission_id": "", # Autofilled
                # "ordering_id": "",
            })

        # Cross-Reference with the other sheet, for a double check of data
        # consistency. Only use the NonPartyTradeNew sheet for now.
        for nonparty_row in row["NonPartyTrade"]:
            pk = party.abbr, period.name, nonparty_row["SubstID"], "NPTImp"
            double_check_old[pk] += decimal.Decimal(nonparty_row["Import"] or 0)
            pk = party.abbr, period.name, nonparty_row["SubstID"], "NPTExp"
            double_check_old[pk] += decimal.Decimal(nonparty_row["Export"] or 0)

        self.double_check(double_check_old, double_check_new, "NonPartyTrade")

        if nonparty_remarks_secretariat:
            nonparty_remarks_secretariat = '\n'.join(nonparty_remarks_secretariat)
        else:
            nonparty_remarks_secretariat = ''
        return nonparty, nonparty_remarks_secretariat

    def double_check(self, old, new, tag):
        """Compare the data from the 'Import' sheet with the data from
        the 'ImportNew' sheet. Log any inconsistencies.
        """
        # Iterate here instead of doing a simple check, so we print out
        # the errors with more details AND to the equals check up to a
        # certain precision because of the floating point operations.
        old = dict(old)
        new = dict(new)
        all_keys = set(list(old.keys()) + list(new.keys()))
        for key in all_keys:
            try:
                old_value = old[key]
            except KeyError:
                logger.warning("%s inconsistency found for %s, present in New but not in base table.",
                               tag, key)
                continue

            try:
                new_value = new[key]
            except KeyError:
                logger.warning("%s inconsistency found for %s, present in base table but not in New.",
                               tag, key)
                continue

            if abs(new_value - old_value) >= (0.1 ** self.precision):
                logger.warning("%s inconsistency found for %s, values differ old=%s new=%s",
                               tag, key, old_value, new_value)
                continue

    def get_exports(self, row, party, period):
        """Parses the export data for the submission identified by
        the party/period combination.

        Returns a list of exports.
        """
        exports = []

        for exports_row in row["Export"]:
            # Get the source party, if not present then add it as NULL
            # the data on the source party will be in the remarks.
            destination_party = exports_row["DestCntryID"].upper()
            if destination_party in ("ZZB", "UNK"):
                destination_party_id = None
            else:
                try:
                    destination_party_id = self.parties[destination_party].id
                except KeyError as e:
                    logger.error("Export unknown source party %s: %s/%s", e, party.abbr,
                                 period.name)
                    destination_party_id = None

            substance = self.substances.get(exports_row["SubstID"])
            substance_id = None
            try:
                substance_id = substance.id
            except AttributeError:
                pass

            blend = None
            blend_id = None
            if not substance:
                blend = self.blends.get(exports_row["SubstID"])
                try:
                    blend_id = blend.id
                except AttributeError:
                    logger.error("Import new unknown substance %s/%s",
                                 party.abbr, period.name)
                    continue

            if not any(exports_row[_npt_type]
                       for _npt_type in ("ExpNew", "ExpRecov")):
                logger.warning("Export no quantity specified: %s/%s/%s", party.abbr,
                               period.name, exports_row["SubstID"])

            critical = False
            if substance and substance.substance_id == 194:
                critical = True

            exports.append({
                "remarks_party": exports_row["Remark"] or "",
                # "remarks_os": "",
                "destination_party_id": destination_party_id,
                "quantity_total_new": exports_row["ExpNew"],
                "quantity_total_recovered": exports_row["ExpRecov"],
                "quantity_feedstock": exports_row["ExpFeedstock"],
                "quantity_critical_uses": exports_row["ExpEssenUse"] if critical else None,
                "quantity_essential_uses": exports_row["ExpEssenUse"] if not critical else None,
                "quantity_high_ambient_temperature": None,
                "quantity_laboratory_analytical_uses": None,
                "quantity_process_agent_uses": exports_row["ExpProcAgent"],
                "quantity_quarantine_pre_shipment": exports_row["ExpQuarAppl"],
                "quantity_polyols": exports_row["ExpPolyol"],
                "quantity_other_uses": None,
                "decision_critical_uses": "",
                "decision_essential_uses": "",
                "decision_high_ambient_temperature": "",
                "decision_laboratory_analytical_uses": "",
                "decision_process_agent_uses": "",
                "decision_quarantine_pre_shipment": "",
                "decision_other_uses": "",
                "blend_id": blend_id if blend else None,
                # "blend_item_id": "",
                "substance_id": substance_id if substance else None,
                # "ordering_id": "",
                # "submission_id": "", # Automatically filled.
            })

        return exports

    def get_destroyed(self, row, party, period):
        """Parses the "destroyed" data for the submission identified by
        the party/period combination.

        Returns a list of destroyed substances.
        """
        destroyed = []

        for destroyed_row in row["Destroy"]:
            substance = self.substances.get(destroyed_row["SubstID"])
            substance_id = None
            try:
                substance_id = substance.id
            except AttributeError:
                pass

            blend = None
            blend_id = None
            if not substance:
                blend = self.blends.get(destroyed_row["SubstID"])
                try:
                    blend_id = blend.id
                except AttributeError:
                    logger.error("Import new unknown substance %s/%s", party.abbr, period.name)
                    continue

            if not destroyed_row["Destroyed"]:
                logger.warning("Destroyed no quantity specified: %s/%s/%s", party.abbr,
                               period.name, destroyed_row["SubstID"])

            destroyed.append({
                "remarks_party": destroyed_row["Remark"] or "",
                "substance_id": substance_id if substance else None,
                # "remarks_os": "",
                "quantity_destroyed": destroyed_row["Destroyed"],
                "blend_id": blend_id if blend else None,
                # "blend_item_id": "",
                # "submission_id": "", # Auto filled
                # "ordering_id": "",
            })

        return destroyed

    def get_produced(self, row, party, period):
        """Parses the "produced" data for the submission identified by
        the party/period combination.

        Returns a list of produced substances.
        """
        produce = []

        for produce_row in row["Produce"]:
            substance = self.substances[produce_row["SubstID"]]
            try:
                substance_id = substance.id
            except KeyError as e:
                logger.error("Produce unknown substance %s: %s/%s", e, party.abbr, period.name)
                continue

            if not produce_row["ProdAllNew"]:
                logger.warning("Produce no quantity specified: %s/%s/%s", party.abbr,
                               period.name, produce_row["SubstID"])

            if substance.substance_id != 194:
                quantity_essen_uses = produce_row["ProdEssenUse"]
                quantity_crit_uses = None
            else:
                quantity_essen_uses = None
                quantity_crit_uses = produce_row["ProdEssenUse"]

            quantity_lab_uses = produce_row["ProdLabUse"]
            decision_lab_uses = ""
            decision_essen_uses = ""
            decision_crit_uses = ""
            remark = produce_row["Remark"] if produce_row["Remark"] is not None else ''
            if quantity_lab_uses and produce_row["ProdEssenUse"]:
                diff = produce_row["ProdEssenUse"] - quantity_lab_uses
                if diff > 0.0001:
                    decision_lab_uses = remark
                    if substance.substance_id != 194:
                        quantity_essen_uses = diff
                        decision_essen_uses = remark
                    else:
                        quantity_crit_uses = diff
                        decision_crit_uses = remark
                elif -0.0001 <= diff <= 0.0001:
                    quantity_lab_uses = quantity_essen_uses if diff >= 0 else quantity_lab_uses
                    quantity_essen_uses = None
                    decision_lab_uses = remark
                else:
                    logger.warning(
                        "Produce: ProdLabUse is greater than ProdEssenUse for %s/%s/%s",
                        party.abbr, period.name, substance.substance_id
                    )

            produce.append({
                "remarks_party": produce_row["Remark"] or "",
                "substance_id": substance_id,
                # "remarks_os": "",
                "quantity_critical_uses": quantity_crit_uses,
                "quantity_essential_uses": quantity_essen_uses,
                "quantity_high_ambient_temperature": None,
                "quantity_laboratory_analytical_uses": quantity_lab_uses,
                "quantity_process_agent_uses": produce_row["ProdProcAgent"],
                "quantity_quarantine_pre_shipment": produce_row["ProdQuarAppl"],
                "quantity_total_produced": produce_row["ProdAllNew"],
                "quantity_other_uses": None,
                "quantity_feedstock": produce_row["ProdFeedstock"],
                "quantity_article_5": produce_row["ProdArt5"],
                "quantity_for_destruction": None,
                "decision_critical_uses": decision_crit_uses,
                "decision_essential_uses": decision_essen_uses,
                "decision_high_ambient_temperature": "",
                "decision_laboratory_analytical_uses": decision_lab_uses,
                "decision_process_agent_uses": "",
                "decision_quarantine_pre_shipment": "",
                "decision_other_uses": "",
                # "submission_id": "", # Auto filled
                # "ordering_id": "",
            })

        return produce

    def get_data(self, row, party, period):
        """Structure and parse the raw data from the Excel file
        so it matches our models.
        """
        # There should only be one entry in the overall sheet.
        try:
            overall = row["Overall"][0]
        except IndexError as e:
            logger.warning("Overall sheet missing for: %s", row)
            return

        # Data is missing these fields, but all submissions seems to have at least one.
        # Fallback for each in order of preference.
        date_reported = overall["DateReported"] or overall["DateCreate"] or overall["DateUpdate"]
        created_at = overall["DateCreate"] or overall["DateReported"] or overall["DateUpdate"]
        updated_at = overall["DateUpdate"] or overall["DateCreate"] or overall["DateReported"]
        date_reported = make_aware(date_reported) if date_reported else None
        created_at = make_aware(created_at) if created_at else None
        updated_at = make_aware(updated_at) if updated_at else None

        if overall["SubmissionType"]:
            try:
                submission_format = SubmissionFormat.objects.get(name=overall["SubmissionType"])
            except SubmissionFormat.DoesNotExist:
                submission_format = None
        else:
            submission_format = None

        if not created_at or not updated_at or not date_reported:
            logger.warning("No date available %s/%s", party.abbr, period.name)

        nonparty, nonparty_remarks_secretariat = self.get_nonparty(
            row, party, period
        )
        return {
            "submission": {
                "schema_version": "legacy",
                "created_at": created_at,
                "updated_at": updated_at,
                "submitted_at": date_reported,
                "version": 1,
                "_workflow_class": "default",
                "_current_state": "finalized",
                "_previous_state": None,
                "flag_provisional": False,
                "flag_valid": True,
                "flag_superseded": False,
                "created_by_id": self.admin.id,
                "last_edited_by_id": self.admin.id,
                "obligation_id": 1,
                "party_id": party.id,
                "reporting_period_id": period.id,
                "cloned_from_id": None,
                # "info_id": "",
                "flag_checked_blanks": bool(overall["Checked_Blanks"]),
                "flag_has_blanks": bool(overall["Blanks"]),
                "flag_confirmed_blanks": bool(overall["Confirm_Blanks"]),
                "flag_has_reported_a1": overall["AI_ComplRep"],
                "flag_has_reported_a2": overall["AII_ComplRep"],
                "flag_has_reported_b1": overall["BI_ComplRep"],
                "flag_has_reported_b2": overall["BII_ComplRep"],
                "flag_has_reported_b3": overall["BIII_ComplRep"],
                "flag_has_reported_c1": overall["CI_ComplRep"],
                "flag_has_reported_c2": overall["CII_ComplRep"],
                "flag_has_reported_c3": overall["CIII_ComplRep"],
                "flag_has_reported_e": overall["EI_ComplRep"],
                "flag_has_reported_f": overall["F_ComplRep"],
                "reporting_channel": ReportingChannel.objects.get(name="Legacy"),
                "questionnaire_remarks_party": "",
                "questionnaire_remarks_secretariat": overall["Remark"] or "",
                "nonparty_remarks_secretariat": nonparty_remarks_secretariat
            },
            "submission_info": {
                "reporting_officer": "",
                "designation": "",
                "organization": "",
                "postal_address": "",
                "country_id": party.id,
                "phone": "",
                "email": "",
                "submission_format": submission_format,
            },
            "art7": {
                "has_imports": overall["Imported"],
                "has_exports": overall["Exported"],
                "has_produced": overall["Produced"],
                "has_destroyed": overall["Destroyed"],
                "has_nonparty": overall["NonPartyTrade"],
                "has_emissions": False,
                # "submission_id": "",
            },
            "imports": self.get_imports(row, party, period),
            "exports": self.get_exports(row, party, period),
            "produced": self.get_produced(row, party, period),
            "destroyed": self.get_destroyed(row, party, period),
            "nonparty": nonparty,
        }

    def check_consistency(self, data, party, period):
        """Check consistency of the data returned by get_data.

        Checks if the data is present the corresponding `has_*` flag
        is correctly set and the other way around.

        Log any inconsistencies, and returns True if the data looks
        to be valid, and False otherwise.
        """
        is_ok = True
        for data_type in self.data_to_check:
            # Check can be done in a single operation, but we want
            # to be verbose to log the inconsistency.
            if data[data_type] and not data["art7"]["has_" + data_type]:
                logger.warning("Inconsistency for %s/%s/%s: has data, but flag is not set",
                               party.abbr, period.name, data_type)
                is_ok = False
            elif not data[data_type] and data["art7"]["has_" + data_type]:
                logger.warning("Inconsistency for %s/%s/%s: does not have data, but flag set",
                               party.abbr, period.name, data_type)
                is_ok = False
        return is_ok

    @transaction.atomic
    def _process_entry(self, party, period, values, recreate=False, purge=False):
        """Inserts the processed data into the DB."""
        is_processed = (party.abbr, period.name) in self.current_submission
        if purge:
            self.delete_instance(party, period)
            return True

        if is_processed:
            if recreate:
                self.delete_instance(party, period)
            else:
                logger.info("Submission %s/%s already imported, skipping.",
                            party.abbr, period.name)
                return False

        submission = Submission.objects.create(
            **values["submission"]
        )

        for key, value in values["submission_info"].items():
            setattr(submission.info, key, value)
        submission.info.save()

        # Use bulk create to bypass any model level validation.
        # This will mean that some entries will be in impossible states but
        # we prefer preserving the legacy data as pristine as possible.

        for key, klass in (
            ("art7", Article7Questionnaire),
            ("imports", Article7Import),
            ("exports", Article7Export),
            ("produced", Article7Production),
            ("destroyed", Article7Destruction),
            ("nonparty", Article7NonPartyTrade),
        ):
            table_values = values[key]
            if isinstance(table_values, list):
                for _i, _instance in enumerate(table_values):
                    klass.objects.create(
                        submission=submission, ordering_id=_i, **_instance
                    )
            else:
                klass.objects.create(
                    submission=submission, **table_values
                )

        # Extra tidy
        submission._current_state = "finalized"
        submission.save()

        # Fill aggregated data on submission import
        submission.fill_aggregated_data()

        if values["submission"]["created_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                created_at=values["submission"]["created_at"]
            )
        if values["submission"]["updated_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                updated_at=values["submission"]["updated_at"]
            )
        for obj in submission.history.all():
            obj.history_user = self.admin
            obj.created_at = values["submission"]["created_at"]
            obj.updated_at = values["submission"]["updated_at"]
            obj.history_date = values["submission"]["created_at"]
            obj.save()

        log_data = ", ".join("%s=%s" % (_data_type, len(values[_data_type]))
                             for _data_type in self.data_to_check)
        logger.info("Submission %s/%s imported with %s",
                    party.abbr, period.name, log_data)
        return True

    def delete_instance(self, party, period):
        """Removes the submission identified by the party and period
        and any related data.
        """
        qs = Submission.objects.filter(
            party=party,
            reporting_period=period,
        ).all()
        for s in qs:
            logger.info("Deleting submission %s/%s", party.abbr, period.name)
            s._current_state = 'data_entry'
            s.save()
            for related_data, aggr_flag in s.RELATED_DATA:
                for instance in getattr(s, related_data).all():
                    logger.debug("Deleting related data: %s", instance)
                    instance.delete()
            s.__class__.data_changes_allowed = True
            s.delete()

    def load_workbook(self, filename, use_cache=False):
        """Loads the Excel file, collating the data based on the
        CntryID and PeriodID.

        If cache is set to True, then the data is loaded from a previously
        loaded version of the file. Reduces time required while doing a lot
        of tests.

        Returns a list in the following format:
        {
            ("RO", 2019): {
                "Overall": [{<overall-data>}],
                "Import": [{<import1>}, {<import2?}, ...]
                ...
            }
            ...
        }
        """
        if use_cache:
            try:
                with open(CACHE_LOC, "rb") as cachef:
                    return pickle.load(cachef)
            except:
                pass

        def get_new():
            return collections.defaultdict(list)

        results = collections.defaultdict(get_new)

        wb = load_workbook(filename=filename)
        for sheet in wb:
            values = list(sheet.values)
            headers = values[0]

            for row in values[1:]:
                row = dict(zip(headers, row))
                if not row["CntryID"]:
                    continue
                pk = row["CntryID"].upper(), row["PeriodID"].upper()
                results[pk][sheet.title].append(row)

        results = list(results.items())
        with open(CACHE_LOC, "wb") as cachef:
            pickle.dump(results, cachef)
        return results

    def handle(self, *args, **options):
        stream = logging.StreamHandler()
        stream.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s %(message)s'
        ))
        logger.addHandler(stream)
        logger.setLevel(logging.WARNING)
        if int(options['verbosity']) > 0:
            logger.setLevel(logging.INFO)
        if int(options['verbosity']) > 1:
            logger.setLevel(logging.DEBUG)

        self.precision = options["precision"]
        single = options["single"]

        try:
            # Create as the first admin we find.
            self.admin = User.objects.filter(is_superuser=True)[0]
        except Exception as e:
            logger.critical("Unable to find an admin: %s", e)
            return

        all_values = self.load_workbook(options["file"], use_cache=options["use_cache"])
        if options['limit']:
            all_values = all_values[:options['limit']]

        success_count = 0
        for pk, values_dict in all_values:
            if single and single != "%s/%s" % pk:
                continue

            logger.debug("Importing row %s", values_dict)

            try:
                party = self.parties[pk[0].upper()]
                period = self.periods[pk[1].upper()]
            except KeyError as e:
                logger.critical("Unable to find matching %s: %s", e, values_dict)
                break

            data = dict(self.get_data(values_dict, party, period))
            self.check_consistency(data, party, period)
            if not options["dry_run"]:
                success_count += self.process_entry(party,
                                                    period,
                                                    data,
                                                    options["recreate"],
                                                    options["purge"])
        logger.info("Success on %s out of %s", success_count, len(all_values))
