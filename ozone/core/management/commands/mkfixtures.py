import json
import os
from copy import deepcopy
from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.core.serializers.json import DjangoJSONEncoder
from django.conf import settings
from openpyxl import load_workbook

from ozone.core.models.exemption import CriticalUseCategory
from ozone.core.models.substance import Blend
from ozone.core.models.utils import RatificationTypes
from ozone.core.models.utils import round_half_up


class Command(BaseCommand):
    help = "Create fixtures from an Excel file"

    MODELS = {
        'annex': {
            'fixture': 'annexes.json',
        },
        'group': {
            'fixture': 'groups.json',
        },
        'language': {
            'fixture': 'languages.json',
        },
        'meeting': {
            'fixture': 'meetings.json',
        },
        'partytype': {
            'fixture': 'partytypes.json'
        },
        'party': {
            'sheet': 'Cntry',
            'fixture': 'parties.json',
        },
        'partyhistory': {
            'fixture': 'partieshistory.json',
            'sheet': 'CntryYr',
        },
        'partyratification': {
            'fixture': 'partiesratification.json',
            'sheet': 'Cntry',
        },
        'mdgregion': {
            'sheet': 'MDG_RegionsCntryArea',
            'fixture': 'mdg_regions.json',
        },
        'region': {
            'sheet': 'Regions',
            'fixture': 'regions.json',
        },
        'subregion': {
            'sheet': 'SubRegion',
            'fixture': 'subregions.json',
        },
        'treaty': {
            'fixture': 'treaties.json',
        },
        'substance': {
            'fixture': 'substances.json',
            'sheet': 'Subst',
        },
        'blend': {
            'fixture': 'blends.json',
            'sheet': 'Blends',
        },
        'blendcomponent': {
            'fixture': 'blend_components.json',
            'sheet': 'BlendComposition',
        },
        'reportingperiod': {
            'fixture': 'reportingperiods.json',
            'sheet': 'Period',
        },
        'substance_edw': {
            'model': 'substance',
            'min_id': 280,
            'fixture': 'substances_edw.json',
            'sheet': 'SubstEDW',
        },
        'baselinetype': {
            'fixture': 'baseline_types.json'
        },
        'baseline': {
            'fixture': 'baselines.json',
            'sheet': 'tbl_prodcons',
            'additional_data': {},
            'prod_transfers': {},
        },
        'criticalusecategory': {
            'fixture': 'critical_use_categories.json',
            'sheet': ['MeBrAgreedCriticalUseCategories', 'MeBrActualCriticalUsebyCategory']
        }
    }
    EXCLUDED = (
        'UNK',
        'CZS',
    )

    OUTPUT_DIR = settings.FIXTURE_DIRS[0]
    FIXTURES = {}

    def add_arguments(self, parser):
        parser.add_argument('file',
                            help="the xlsx input file")
        parser.add_argument('models', nargs='+',
                            help="a list of meta-models to generate, e.g. party, substance")

    def handle(self, *args, **options):
        wb = load_workbook(filename=options['file'])
        print('Excel file loaded.')

        # preload existing fixtures for dependent models
        for model in self.MODELS.keys():
            try:
                filename = os.path.join(
                    self.OUTPUT_DIR, self.MODELS[model]['fixture'])
                with open(filename, 'r', encoding='utf-8') as f:
                    self.FIXTURES[model] = json.load(f)
            except FileNotFoundError:
                continue

        for model in options['models']:
            data = self.parse_model(wb, model)
            filename = os.path.join(
                self.OUTPUT_DIR, self.MODELS[model]['fixture'])
            with open(filename, 'w', encoding="utf-8") as outfile:
                json.dump(
                    data, outfile,
                    indent=2, ensure_ascii=False, sort_keys=True,
                    cls=DjangoJSONEncoder
                )
                print('Done with %s' % filename)

    def row2dict(self, sheet, row, index):
        header = [cell.value for cell in sheet[1]]
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        values['_index'] = index
        return values

    def lookup_id(self, model, field, key, data=None):
        # This method used to be pretty until UNK SubRegion appeared several
        # times
        if not key:
            return None
        if not data:
            data = self.FIXTURES[model]
        objs = list(filter(lambda x: x['fields'][field] == key, data))
        if len(objs) == 0:
            raise CommandError('Cannot find "{}" having "{}" = "{}"'.format(
                model, field, key
            ))
        if len(objs) == 1:
            return objs[0]['pk']
        return objs

    def parse_model(self, workbook, model):
        # reinitialize current model in FIXTURES
        data = self.FIXTURES[model] = []
        if not self.MODELS.get(model) or not self.MODELS[model].get('sheet'):
            raise CommandError(
                'Import for model "%s" is not implemented' % model)
        sheet_names = self.MODELS[model]['sheet']
        if type(sheet_names) == str:
            # To support multiple sheets
            sheet_names = [sheet_names]

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            for idx in range(1, sheet.max_row):
                row = self.row2dict(sheet, sheet[idx + 1], idx)
                model_class = 'core.' + (self.MODELS[model].get('model') or model)
                obj = {
                    'pk': len(data) + 1 + (self.MODELS[model].get('min_id') or 0),
                    'model': model_class,
                    'fields': {},
                }
                fields = getattr(self, model + "_map")(obj['fields'], row)
                if isinstance(fields, dict):
                    obj['fields'] = fields
                    data.append(obj)
                elif isinstance(fields, list):
                    # multiple objects returned
                    for idx, f in enumerate(fields):
                        data.append({
                            'pk': obj['pk'] + idx,
                            'model': obj['model'],
                            'fields': f,
                        })
                    pass
                else:
                    # Probably None was returned, don't add to list
                    continue

        if hasattr(self, model + "_additional_data"):
            data += getattr(self, model + "_additional_data")(len(data) + 1)

        # if a post process method exists, invoke it
        if hasattr(self, model + "_postprocess"):
            for obj in data:
                getattr(self, model + "_postprocess")(obj['fields'])

        if hasattr(self, model + "_aggregate"):
            data = getattr(self, model + "_aggregate")(data)
        return data

    def mdgregion_map(self, f, row):
        f['code'] = row['M49Codes']
        f['name'] = row['RegionCountryArea']
        f['income_type'] = row['IncomeGroup']
        f['remark'] = row['Remark'] or ''
        if 'parent_regions' not in f:
            f['parent_regions'] = list()
        if row['Group_Code'] != '-1':
            f['parent_regions'].append(row['Group_Code'])
        return f

    def mdgregion_aggregate(self, data):
        # pass 1: join the regions with the same code
        newdata = {}
        for obj in data:
            # use code as pk
            code = obj['fields']['code']
            if code in newdata:
                newdata[code]['fields']['parent_regions'] += (
                    obj['fields']['parent_regions']
                )
            else:
                newdata[code] = deepcopy(obj)
                # override the pk
                newdata[code]['pk'] = code

        return list(newdata.values())

    def region_map(self, f, row):
        f['name'] = row['RegionName']
        f['abbr'] = row['RegionID']
        f['remark'] = row['Remark'] or ''
        # unused: RegionNameFr, RegionNameSp
        return f

    def subregion_map(self, f, row):
        f['name'] = row['SubRegionName']
        f['abbr'] = row['SubRegionID']
        f['region'] = self.lookup_id('region', 'abbr', row['RegionID'])
        f['remark'] = row['Remark'] or ''
        return f

    def party_map(self, f, row):
        f['abbr'] = row['CntryID']
        if row['CntryID'][:2] == 'ZZ' or row['CntryID'].upper() in self.EXCLUDED:
            # Remove "All countries" and "Some countries"
            # f['_deleted'] = True
            return None
        if row['CntryName'] == 'Taiwan Province':
            f['name'] = 'Taiwan, Province of China'
        else:
            f['name'] = row['CntryName']
        # Skipped fields: CntryNameFr, CntryNameSp, PrgApprDate, CntryID_org,
        # CntryName20, MDG_CntryCode, ISO_Alpha3Code, www_country_id

        # RegionID is not necessary because it's implied from SubRegionID
        subregion = self.lookup_id('subregion', 'abbr', row['SubRegionID'])
        if isinstance(subregion, list):
            # UNK subregion is not unique, so lookup_id will return a list
            region = self.lookup_id('region', 'abbr', row['RegionID'])
            subregion = self.lookup_id(
                'subregion', 'region', region, subregion)
        f['subregion'] = subregion

        f['remark'] = row['Remark'] or ""
        # parent_party will be replaced in party_postprocess
        f['parent_party'] = row['MainCntryID']
        f['iso_alpha3_code'] = row['ISO_Alpha3Code'] or ''
        f['abbr_alt'] = row['CntryID_org'] or ''
        f['name_alt'] = row['CntryName20'] or ''
        try:
            f['mdg_region'] = self.lookup_id('mdgregion', 'code', str(row['MDG_CntryCode']))
        except CommandError:
            # MDG region for Antarctica does not exist
            pass
        return f

    def party_postprocess(self, f):
        # set parent party by looking up in the same data
        if 'parent_party' in f:
            f['parent_party'] = self.lookup_id(
                'party', 'abbr', f['parent_party'])

    def partyhistory_map(self, f, row):
        if row['CntryID'] == 'HOLV' or row['CntryID'].upper() in self.EXCLUDED:
            # f['party'] = self.lookup_id('party', 'abbr', 'VA')
            # f['_deleted'] = True
            return None
        else:
            f['party'] = self.lookup_id('party', 'abbr', row['CntryID'])
        f['reporting_period'] = self.lookup_id(
            'reportingperiod', 'name', row['PeriodID'])
        f['population'] = row['Population'] if row['Population'] else 0
        art5_group2 = ["BH", "IN", "IR", "IQ",
                       "KW", "OM", "PK", "QA", "SA", "AE"]
        non_art5_group2 = ["BY", "KZ", "RU", "TJ", "UZ"]
        period_datetime = datetime.strptime(
            self.FIXTURES['reportingperiod'][
                f['reporting_period'] - 1]['fields']['end_date'], "%Y-%m-%d"
        )
        if period_datetime < datetime.strptime('2019-01-01', "%Y-%m-%d"):
            if row['Article5']:
                f['party_type'] = self.lookup_id(
                    'partytype', 'name', 'Article 5'
                )
                f['is_article5'] = True
            else:
                f['party_type'] = self.lookup_id(
                    'partytype', 'name', 'Non Article 5'
                )
                f['is_article5'] = False
        else:
            if row['Article5']:
                if row['CntryID'] in art5_group2:
                    f['party_type'] = self.lookup_id(
                        'partytype', 'name', 'Article 5 Group 2'
                    )
                else:
                    f['party_type'] = self.lookup_id(
                        'partytype', 'name', 'Article 5 Group 1'
                    )
                f['is_article5'] = True
            else:
                if row['CntryID'] in non_art5_group2:
                    f['party_type'] = self.lookup_id(
                        'partytype', 'name', 'Non Article 5 Group 2'
                    )
                else:
                    f['party_type'] = self.lookup_id(
                        'partytype', 'name', 'Non Article 5 Group 1'
                    )
                f['is_article5'] = False

        hat_parties = [
            "DZ", "BH", "BJ", "BF", "CF", "TD", "CI", "DJ", "EG", "ER", "GM",
            "GH", "GN", "GW", "IR", "IQ", "JO", "KW", "LY", "ML", "MR", "NE",
            "NG", "OM", "PK", "QA", "SA", "SN", "SD", "SY", "TG", "TN", "TM", "AE"
        ]
        if row['CntryID'] in hat_parties:
            f['is_high_ambient_temperature'] = True
        else:
            f['is_high_ambient_temperature'] = False
        f['is_eu_member'] = row['EurUnion']
        f['is_ceit'] = row['CEIT']
        f['remark'] = row['Remark'] if row['Remark'] else ""
        return f

    def partyratification_map(self, f, row):
        if row['CntryID'][:2] == 'ZZ' or row['CntryID'] != row['MainCntryID'] \
                or row['CntryID'].upper() in self.EXCLUDED:
            # Remove "All countries" and "Some countries"
            return None
        ratif_types_map = {
            'Ac': RatificationTypes.ACCESSION.value,
            'Ap': RatificationTypes.APPROVAL.value,
            'At': RatificationTypes.ACCEPTANCE.value,
            'R': RatificationTypes.RATIFICATION.value,
            'Sc': RatificationTypes.SUCCESSION.value,
            # TODO: What about RatificationTypes.SIGNING ??
        }
        objs = []
        party = self.lookup_id('party', 'abbr', row['CntryID'])
        treaties = ['VC', 'MP', 'LA', 'CA', 'MA', 'BA', 'KA']
        for treaty_id in treaties:
            if row['RD_' + treaty_id]:
                ratification_date = row['RD_' + treaty_id].date()
                entry_into_force_datetime = row['EIF_' + treaty_id]
                entry_into_force_date = entry_into_force_datetime.date() if entry_into_force_datetime else None
                objs.append({
                    'party': party,
                    'treaty': self.lookup_id('treaty', 'treaty_id', treaty_id),
                    'ratification_type': ratif_types_map.get(row['RT_' + treaty_id], ""),
                    'ratification_date': ratification_date,
                    'entry_into_force_date': entry_into_force_date,
                })
        return objs

    def substance_map(self, f, row):
        f['substance_id'] = row['SubstID']
        f['name'] = row['SubstName']
        # Skipped fields: SubstNameFr, SubstNameSp
        annex = row['Anx']
        is_captured = False
        group = None
        if annex and annex != '-':
            group = annex + row['Grp']
            if f['name'] == 'HFC-23':
                group = 'F'
                is_captured = True
            if group == 'FI' or group == 'FII':
                group = 'F'
            if group == 'FIII':
                # Most of these are blends. Those which are not were moved to SubstEDW
                return
            f['group'] = self.lookup_id('group', 'group_id', group)

        sort_order_extra = {
            'AI': 0,
            'AII': 10000,
            'BI': 20000,
            'BII': 30000,
            'BIII': 40000,
            'CI': 50000,
            'CII': 60000,
            'CIII': 70000,
            'EI': 80000,
            'F': 90000,
            'X': 100000,
        }
        # Make sort_order global (for all groups)
        f['sort_order'] = (
            (row['AnxGrpSort'] or 9999) +
            sort_order_extra[group or 'X']
        )
        f['odp'] = row['SubstODP']
        f['gwp'] = row['SubstGWP']
        f['max_odp'] = row['MaxODP'] or 0
        f['min_odp'] = row['MinODP'] or 0
        f['description'] = row['SubstDescr'] or row['SubstName']
        f['formula'] = row['SubstFormula'] or ""
        f['number_of_isomers'] = row['Number of Isomers']
        f['gwp2'] = row['GWP']
        f['gwp_error_plus_minus'] = row['GWP_Error_Plus_Minus']
        f['remark'] = row['Remark'] or ""
        f['carbons'] = row['Carbons'] or ""
        f['hydrogens'] = row['Hydrogens'] or ""
        f['fluorines'] = row['Fluorines'] or ""
        f['chlorines'] = row['Chlorines'] or ""
        f['bromines'] = row['Bromines'] or ""
        f['is_contained_in_polyols'] = f['name'] in ['CFC-11', 'HCFC-141B']
        f['is_captured'] = is_captured
        f['has_critical_uses'] = (f['name'] == 'Methyl Bromide')

        # TODO: Not mapped: r_code, mp_control, main_usage
        return f

    def substance_edw_map(self, f, row):
        # F/III substances which are not blends and should not be skipped
        row['Anx'] = None
        row['Grp'] = None
        return self.substance_map(f, row)

    def blend_map(self, f, row):
        f['blend_id'] = row['Blend']
        f['legacy_blend_id'] = row['BlendID']
        # There is a new column called BlendType, but this mapping also works fine
        if f['blend_id'].startswith('R-4'):
            f['type'] = Blend.BlendTypes.ZEOTROPE.value
            f['sort_order'] = 110000 + row['_index']
        elif f['blend_id'].startswith('R-5'):
            f['type'] = Blend.BlendTypes.AZEOTROPE.value
            f['sort_order'] = 120000 + row['_index']
        elif f['blend_id'].startswith('MeBr'):
            f['type'] = Blend.BlendTypes.MeBr.value
            f['sort_order'] = 140000 + row['_index']
        else:
            f['type'] = Blend.BlendTypes.OTHER.value
            f['sort_order'] = 130000 + row['_index']

        f['composition'] = row['Composition']
        f['other_names'] = row['OtherNames'] or ""
        f['remark'] = row['Remark'] or ""

        f['odp'] = row['BlendODP']
        f['gwp'] = row['BlendGWP']
        f['trade_name'] = row['BlendTradeName'] or ""
        f['composition_alt'] = row['CompositionOrg'] or ""
        f['cnumber'] = row['CNumber'] or ""

        # TODO: Not mapped: main_usage, hcfc, hfc, mp_control
        return f

    def blendcomponent_map(self, f, row):
        f['blend_id'] = self.lookup_id('blend', 'blend_id', row['Blend'])
        f['component_name'] = row['Component'] or ""
        f['percentage'] = row['Percentage']
        f['cnumber'] = row['CNumber'] or ""
        try:
            substance = self.lookup_id('substance', 'substance_id', row['SubstID'])
        except CommandError:
            substance = self.lookup_id(
                'substance_edw', 'substance_id', row['SubstID'])
        f['substance'] = substance
        return f

    def reportingperiod_map(self, f, row):
        f['name'] = row['PeriodID']
        f['start_date'] = row['StartDate'].date()
        f['end_date'] = row['EndDate'].date()
        f['description'] = row['PeriodDescr']
        f['is_reporting_allowed'] = f['name'][0] == 'C' or (
            f['name'].isdigit() and int(f['name']) <= 2018
        )
        f['is_reporting_open'] = f['name'] in ('2017', '2018')
        return f

    def reportingperiod_additional_data(self, idx):
        objs = [
            {
                "fields": {
                    "description": "",
                    "end_date": datetime.strptime("1987-12-31", "%Y-%m-%d").date(),
                    "is_reporting_allowed": False,
                    "is_reporting_open": False,
                    "name": "1987",
                    "start_date": datetime.strptime("1987-01-01", "%Y-%m-%d").date()
                },
                "model": "core.reportingperiod",
                "pk": idx
            },
            {
                "fields": {
                    "description": "",
                    "end_date": datetime.strptime("1988-12-31", "%Y-%m-%d").date(),
                    "is_reporting_allowed": False,
                    "is_reporting_open": False,
                    "name": "1988",
                    "start_date": datetime.strptime("1988-01-01", "%Y-%m-%d").date()
                },
                "model": "core.reportingperiod",
                "pk": idx + 1
            }
        ]
        return objs

    def baseline_map(self, f, row):
        additional_periods = ['1995', '1996', '1997', '1998', '1999', '2000']
        baseline_periods = ['BaseA5', 'BaseNA5']
        bdn_groups = ['AI', 'AII', 'BI', 'EI']
        group = row['Anx'] + row['Grp']
        if row['PeriodID'] in additional_periods and group in bdn_groups:
            if not self.MODELS['baseline']['additional_data'].get(row['CntryID']):
                self.MODELS['baseline']['additional_data'][row['CntryID']] = {}
                for bdn_group in bdn_groups:
                    self.MODELS['baseline']['additional_data'][row['CntryID']][bdn_group] = {}
            self.store_baseline_additional_data(
                row['CntryID'],
                row['PeriodID'],
                group,
                row['ProdArt5'] if row['ProdArt5'] else 0,
            )
        elif row['PeriodID'] in baseline_periods:
            if row['PeriodID'] == 'BaseNA5' and row['ProdTransfer']:
                self.store_baseline_prod_transfers(
                    row['CntryID'],
                    group,
                    row['ProdTransfer']
                )
            entries = []
            party = self.lookup_id('party', 'abbr', row['CntryID'])
            group = self.lookup_id('group', 'group_id', group)
            party_type = row['PeriodID'][4:]

            f['party'] = party
            f['group'] = group
            f['baseline_type_id'] = self.lookup_id(
                'baselinetype',
                'name',
                party_type + 'Prod'
            )
            # Don't need to call get_decimals for 'BaseA5' and 'BaseNA5' periods
            # because is not a special case and we will round to 1 decimal.
            baseline_prod = row['CalcProd']
            if baseline_prod:
                baseline_prod = round_half_up(row['CalcProd'], 1)
            f['baseline'] = baseline_prod
            entries.append(f)

            f = {}
            f['party'] = party
            f['group'] = group
            f['baseline_type_id'] = self.lookup_id(
                'baselinetype',
                'name',
                party_type + 'Cons'
            )
            baseline_cons = row['CalcCons']
            if baseline_cons:
                baseline_cons = round_half_up(row['CalcCons'], 1)
            f['baseline'] = baseline_cons
            entries.append(f)

            print('Procces country {} for group {}'.format(row['CntryID'], row['Anx'] + row['Grp']))

            return entries
        else:
            return

    def store_baseline_additional_data(self, party, period, group, prod):
        """
        Returns a dictionary in the following format:
        {
            'CntryID':  {
                'GroupID': {
                    '1995': 'ProdArt5': val1,
                    '1996': 'ProdArt5': val2,
                    ...
                    '2000': 'ProdArt5': val3
                }
            }
        }
        """

        self.MODELS['baseline']['additional_data'][party][group][period] = {
            'ProdArt5': prod,
        }

    def store_baseline_prod_transfers(self, party, group, prod_transfer):
        self.MODELS['baseline']['prod_transfers'][(party, group)] = prod_transfer

    def baseline_additional_data(self, idx):
        entries = []
        bdn_groups = ['AI', 'AII', 'BI', 'EI']
        for party in self.MODELS['baseline']['additional_data'].keys():
            for group in bdn_groups:
                obj = {}
                obj['model'] = "core.baseline"
                obj['fields'] = {}
                obj['fields']['party'] = self.lookup_id('party', 'abbr', party)
                obj['fields']['group'] = self.lookup_id('group', 'group_id', group)
                obj['fields']['baseline_type_id'] = self.lookup_id('baselinetype', 'name', 'BDN_NA5')
                data = self.MODELS['baseline']['additional_data'][party][group]
                data['prod_transfers'] = self.MODELS['baseline']['prod_transfers'].get((party, group))
                obj['fields']['baseline'] = getattr(self, 'get_bdn_' + group)(data)
                obj['pk'] = idx
                entries.append(obj)
                idx += 1
                print('Procces additional data country {} for group {}'.format(party, group))

        return entries

    def get_bdn_AI(self, data):
        periods = ['1995', '1996', '1997']
        return self.calc_avg(data, periods)

    def get_bdn_AII(self, data):
        periods = ['1995', '1996', '1997']
        return self.calc_avg(data, periods)

    def get_bdn_BI(self, data):
        periods = ['1998', '1999', '2000']
        return self.calc_avg(data, periods)

    def get_bdn_EI(self, data):
        periods = ['1995', '1996', '1997', '1998']
        return self.calc_avg(data, periods)

    def calc_avg(self, data, periods):
        s = 0
        for period in periods:
            if not data.get(period):
                return
            s += data[period]['ProdArt5']
        baseline = s / len(periods)
        if data['prod_transfers']:
            baseline += data['prod_transfers']

        # BDN new baselines are rounded to 5 decimals
        return round_half_up(baseline, 5)

    def criticalusecategory_map(self, f, row):
        columns = [
            'Categories of permitted critical uses',
            'CU_Title',
        ]
        for col in columns:
            if col in row:
                f['name'] = row[col]
        return f

    def criticalusecategory_aggregate(self, data):
        # Cleanup some duplicates
        data.sort(key=lambda x: x['fields']['name'].strip().lower())
        # sort first just to get rid of some derived names
        newdata = {}
        for obj in data:
            name = obj['fields']['name']
            key = CriticalUseCategory.get_alt_name(name)
            if key not in newdata:
                name = name.strip()
                obj['fields']['name'] = name[0].upper() + name[1:]
                obj['fields']['code'] = key
                obj['pk'] = len(newdata) + 1
                newdata[key] = obj
            # else continue

        return sorted(
            list(newdata.values()),
            key=lambda x: x['pk']
        )
