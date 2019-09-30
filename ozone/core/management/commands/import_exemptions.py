import logging
from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.utils.timezone import make_aware

from ozone.core.models import (
    Party,
    Substance,
    Submission,
    ReportingPeriod,
    ReportingChannel,
    RAFReport,
    RAFReportUseCategory,
    RAFImport,
    Obligation,
    Nomination,
    ExemptionApproved,
    CriticalUseCategory,
    ApprovedCriticalUse,
)
from ozone.core.models.utils import (
    float_to_decimal,
    float_to_decimal_zero_if_none
)


User = get_user_model()


logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = "Import two types of submissions: submission for obligation 'Exemption' and submission for obligation 'Essential and Critical uses (RAF)'."

    def __init__(self, stdout=None, stderr=None, no_color=False):
        super().__init__(stdout=None, stderr=None, no_color=False)

        self.periods = {_period.name: _period
                        for _period in ReportingPeriod.objects.all()}
        self.parties = {
            _party.abbr: _party
            for _party in Party.objects.all()
        }
        # Append legacy EU code
        self.parties.update({
            'ECE': Party.objects.get(abbr='EU')
        })
        self.substances = {_substance.substance_id: _substance
                           for _substance in Substance.objects.all()}

    def add_arguments(self, parser):
        parser.add_argument('file',
                            help="the xlsx input file")
        parser.add_argument('--recreate', action="store_true", default=False,
                            help="Re-create if submission already exists.")
        parser.add_argument('--purge', action="store_true", default=False,
                            help="Purge all entries that were imported")
        parser.add_argument("--dry-run", action="store_true", default=False,
                            help="Only parse the data, but do not insert it.")

    def handle(self, *args, **options):
        stream = logging.StreamHandler()
        stream.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s %(message)s'
        ))
        logger.addHandler(stream)
        logger.setLevel(logging.WARNING)
        if int(options['verbosity']) > 0:
            logger.setLevel(logging.INFO)
        if int(options['verbosity']) > 1:
            logger.setLevel(logging.DEBUG)

        try:
            # Create as the first admin we find.
            self.admin = User.objects.filter(is_superuser=True)[0]
        except Exception as e:
            logger.critical("Unable to find an admin: %s", e)
            return

        raf_values = self.load_raf_sheet(options["file"])
        exemption_values = self.load_exemption_sheets(options["file"])

        if not options["dry_run"]:
            self.import_data(raf_values, Obligation.objects.get(pk=2), options)
            self.import_data(exemption_values, Obligation.objects.get(pk=11), options)

    def load_raf_sheet(self, filename):
        """
        Loads the 'EssenUse' from the Excel file.
        Returns a dictionary in the following format:
        {
            ("CntryID", PeriodID): {
                (
                    "SubstID",
                    "EssenCrit"
                ):
                    [ { data1 }, { data2 } ]
                ....
            },
            ...
        }
        where
        * [{data_i}] is a list of dictionaries where keys are the headers
        of the sheet and values are the values from row number i.
        """

        results = {}
        # Dictionary used for avoiding duplicates.
        avoid_duplicates = {}

        wb = load_workbook(filename=filename)
        ws = wb['EssenUse']

        values = list(ws.values)
        headers = values[0]
        for row in values[1:]:
            row = dict(zip(headers, row))
            pk1 = row["CntryID"], row["PeriodID"]
            if not results.get(pk1):
                results[pk1] = {}
            pk2 = (
                row["SubstID"],
                row["ExemptType"] == "EmergCrit" if row["ExemptType"] else False,
            )

            avoid_duplicates_pk = (
                row["CntryID"],
                row["PeriodID"],
                row["SubstID"],
                row["ImpSrcCntryID"],
            )
            if not results[pk1].get(pk2):
                results[pk1][pk2] = {"raf": [], "critical": []}
                results[pk1][pk2]["raf"].append(row)
                avoid_duplicates[avoid_duplicates_pk] = True
            else:
                if avoid_duplicates.get(avoid_duplicates_pk):
                    # TODO: what is going on here?
                    if row["Exempted"]:
                        for existing_row in results[pk1][pk2]["raf"]:
                            if existing_row["ImpSrcCntryID"] == row["ImpSrcCntryID"]:
                                existing_row["Exempted"] = row["Exempted"]
                else:
                    results[pk1][pk2]["raf"].append(row)
                    avoid_duplicates[avoid_duplicates_pk] = True

        ws = wb['MeBrActualCriticalUsebyCategory']
        values = list(ws.values)
        headers = values[0]
        for row in values[1:]:
            row = dict(zip(headers, row))
            pk1 = row["CntryID"], row["PeriodID"]
            if not results.get(pk1):
                results[pk1] = {}
            pk2 = (
                Substance.objects.get(name='Methyl Bromide').substance_id, False
            )
            if not results[pk1].get(pk2):
                results[pk1][pk2] = {"raf": [], "critical": []}
            results[pk1][pk2]["critical"].append(row)

        return results

    def load_exemption_sheets(self, filename):
        """
        Loads the 'EssenNom', 'EssenExemp' and 'MeBrAgreedCriticalUseCategories'.
        Returns a dictionary in the following format:
        {
            ("CntryID", PeriodID): {
                "EssenNom': [ { data1 }, { data2 }, ... ],
                "EssenExemp": [ { data1 }, { data2 }, ... ]
            }
        }
        where
        * [{data_i}] is a list of dictionaries where keys are the headers
        of the sheet (EssenNom or EssenExemp) and values are the values from
        row number i.
        """

        results = {}

        wb = load_workbook(filename=filename)

        ws_list = [wb['EssenNom'], wb['EssenExemp'], wb['MeBrAgreedCriticalUseCategories']]

        for sheet in ws_list:
            values = list(sheet.values)
            headers = values[0]
            for row in values[1:]:
                row = dict(zip(headers, row))
                pk = row["CntryID"], row["PeriodID"]
                if not results.get(pk):
                    results[pk] = {}
                    # Initialize both lists.
                    results[pk]['EssenNom'] = []
                    results[pk]['EssenExemp'] = []
                    results[pk]['MeBrAgreedCriticalUseCategories'] = []
                    results[pk][sheet.title].append(row)
                else:
                    results[pk][sheet.title].append(row)

        return results

    def import_data(self, values, obligation, options):
        success_count = 0
        for (party_abbr, period_name), rows in values.items():
            success_count += self.process_entry(
                party_abbr,
                period_name,
                rows,
                obligation,
                recreate=options["recreate"],
                purge=options["purge"]
            )

        logger.info("Success on %s out of %s", success_count, len(values))

    def process_entry(self, party_abbr, period_name, rows, obligation, recreate=False, purge=False):
        """
        Process the parsed data and insert it into the DB.
        Only a wrapper, see process_submission_exemption_entry and process_submission_raf_entry.
        """

        try:
            if obligation.pk == 2:
                return self.process_submission_raf_entry(
                    party_abbr, period_name, rows, obligation, recreate=recreate, purge=purge
                )
            elif obligation.pk == 11:
                return self.process_submission_exemption_entry(
                    party_abbr, period_name, rows, obligation, recreate=recreate, purge=purge
                )
        except KeyboardInterrupt:
            raise
        except Exception as e:
            logger.error("Error %s while saving: %s/%s", e, party_abbr, period_name,
                         exc_info=True)
            return False

    def process_submission_raf_entry(self, party_abbr, period_name, rows, obligation, recreate=False, purge=False):
        """
        Inserts the processed data into the DB.
        """
        logger.info(
            f'{"Deleting" if purge else "Creating"} RAF for '
            f'{party_abbr}/{period_name}'
        )

        try:
            party = self.parties[party_abbr]
            period = self.periods[period_name]
        except KeyError as e:
            logger.critical("Unable to find matching %s: %s", e, rows)
            return False

        if not period.is_reporting_allowed:
            logger.error(
                "Error `Reporting cannot be performed for this reporting period.`"
                "while saving: %s/%s ",
                party_abbr,
                period_name,
                exc_info=True
            )
            return False

        if purge:
            self.delete_instance(party, period, obligation)
            return True

        if recreate:
            self.delete_instance(party, period, obligation)

        data = self.get_submission_raf_data(party, period, rows)

        submission = Submission.objects.create(
            **data["submission"]
        )

        for key, value in data["submission_info"].items():
            setattr(submission.info, key, value)
        submission.info.save()

        for raf in data['rafreports']:
            # Create RAFReport objects
            report = RAFReport.objects.create(
                submission=submission,
                **raf['raf_report']
            )

            # Create RAFImport objects, if they exist.
            for raf_import in raf['imports']:
                RAFImport.objects.create(
                    report=report,
                    **raf_import
                )

            # Create critical use category objects, if they exist
            for raf_critical_use in raf['use_categories']:
                RAFReportUseCategory.objects.create(
                    report=report,
                    **raf_critical_use
                )

        submission._current_state = "finalized"
        submission.save()

        # Setting updated_at and created_at like this avoids creating a new
        # history item.
        if data["submission"]["created_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                created_at=data["submission"]["created_at"]
            )
        if data["submission"]["updated_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                updated_at=data["submission"]["updated_at"]
            )
        for obj in submission.history.all():
            obj.history_user = self.admin
            if data["submission"]["created_at"]:
                obj.history_date = data["submission"]["created_at"]
                obj.created_at = data["submission"]["created_at"]
            if data["submission"]["updated_at"]:
                obj.updated_at = data["submission"]["updated_at"]
            obj.save()

        return True

    def process_submission_exemption_entry(self, party_abbr, period_name, rows, obligation, recreate=False, purge=False):
        """
        Inserts the processed data into the DB.
        """

        try:
            party = self.parties[party_abbr]
            period = self.periods[period_name]
        except KeyError as e:
            logger.critical("Unable to find matching %s: %s", e, rows)
            return False

        if not period.is_reporting_allowed:
            logger.error(
                "Error `Reporting cannot be performed for this reporting period.` "
                "while saving: %s/%s ",
                party_abbr,
                period_name,
                exc_info=True
            )
            return False

        if purge:
            self.delete_instance(party, period, obligation)
            return True

        if recreate:
            self.delete_instance(party, period, obligation)

        data = self.get_submission_exemption_data(party, period, rows)

        logger.info(f'Creating Exemption for {party_abbr}/{period_name}')
        submission = Submission.objects.create(
            **data["submission"]
        )

        for key, value in data["submission_info"].items():
            setattr(submission.info, key, value)
        submission.info.save()

        for key, klass in (
            ("nominations", Nomination),
            ("exemptionapproveds", ExemptionApproved),
        ):
            for val in data[key]:
                klass.objects.create(
                    submission=submission,
                    **val
                )
        self.create_approved_critical_uses(
            party, period, data['approved_critical_uses']
        )

        submission._current_state = "finalized"
        submission.save()

        # Setting updated_at and created_at like this avoids creating a new
        # history item.
        if data["submission"]["created_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                created_at=data["submission"]["created_at"]
            )
        if data["submission"]["updated_at"]:
            Submission.objects.filter(pk=submission.pk).update(
                updated_at=data["submission"]["updated_at"]
            )
        for obj in submission.history.all():
            obj.history_user = self.admin
            if data["submission"]["created_at"]:
                obj.created_at = data["submission"]["created_at"]
                obj.history_date = data["submission"]["created_at"]
            if data["submission"]["updated_at"]:
                obj.updated_at = data["submission"]["updated_at"]
            obj.save()

        return True

    def delete_instance(self, party, period, obligation):
        """
        Removes the submission identified by the party, period and obligation
        and any related data.
        """

        qs = Submission.objects.filter(
            party=party,
            reporting_period=period,
            obligation=obligation
        ).all()

        for sub in qs:
            logger.info("Deleting submission %s/%s", party.abbr, period.name)
            sub._current_state = 'data_entry'
            sub.save()
            for related_data, aggr_flag in sub.RELATED_DATA:
                for instance in getattr(sub, related_data).all():
                    logger.debug("Deleting related data: %s", instance)
                    instance.delete()
            sub.__class__.data_changes_allowed = True
            sub.delete()

    def get_submission_raf_data(self, party, period, rows):
        """
        Structure and parse the raw data from the Excel file
        so it matches our Submission model with RAF obligation.
        """

        # Because there are multiple 'DateCreate' values we will take the latest value.
        created_at = None
        for pk in rows.keys():
            # pk = ("SubstID", "Produced", "OpenBal", "EssenUse", "Exported", "Destroyed", "EssenCrit")
            for row in rows[pk]["raf"]:
                if not created_at and row["DateCreate"]:
                    created_at = row["DateCreate"]
                elif created_at and row["DateCreate"]:
                    if row["DateCreate"] < created_at:
                        created_at = row["DateCreate"]

        # # Similar to `DateCreate` but this time we will take the newest value.
        updated_at = None
        for pk in rows.keys():
            for row in rows[pk]["raf"]:
                if not updated_at and row["DateUpdate"]:
                    updated_at = row["DateUpdate"]
                elif updated_at and row["DateUpdate"]:
                    if row["DateUpdate"] < updated_at:
                        updated_at = row["DateUpdate"]

        if not created_at:
            created_at = updated_at
        elif not updated_at:
            updated_at = created_at

        created_at = make_aware(created_at) if created_at else created_at
        updated_at = make_aware(updated_at) if updated_at else updated_at

        return {
            "submission": {
                "schema_version": "legacy",
                "created_at": created_at,
                "updated_at": updated_at,
                "version": 1,
                "_workflow_class": "default",
                "_current_state": "finalized",
                "_previous_state": None,
                "flag_provisional": False,
                "flag_valid": True,
                "flag_superseded": False,
                "created_by_id": self.admin.id,
                "last_edited_by_id": self.admin.id,
                "obligation_id": 2,
                "party_id": party.id,
                "reporting_period_id": period.id,
                "cloned_from_id": None,
                # "info_id": "",
                "reporting_channel": ReportingChannel.objects.get(name="Legacy")
            },
            "submission_info": {
                "reporting_officer": "",
                "designation": "",
                "organization": "",
                "postal_address": "",
                "country_id": party.id,
                "phone": "",
                "email": "",
                "date": created_at
            },
            "rafreports": self.get_rafs(party, period, rows),
        }

    def get_submission_exemption_data(self, party, period, rows):
        """
        Structure and parse the raw data from the Excel file
        so it matches our Submission model with Exemption obligation.
        """

        # Because there are multiple 'DateCreate' and 'SubmitDate' values
        # we will take the latest value.
        date_seek_rows = rows['EssenNom'] + rows['EssenExemp']
        submitted_at_list = [entry['SubmitDate'] for entry in date_seek_rows if entry.get('SubmitDate')]
        submitted_at = min(submitted_at_list) if submitted_at_list else None
        updated_at_list = [entry['DateUpdate'] for entry in date_seek_rows if entry.get('DateUpdate')]
        updated_at = min(updated_at_list) if updated_at_list else None
        created_at_list = [entry['DateCreate'] for entry in date_seek_rows if entry.get('DateCreate')]
        created_at = min(created_at_list) if created_at_list else updated_at

        submitted_at = make_aware(submitted_at) if submitted_at else submitted_at
        created_at = make_aware(created_at) if created_at else created_at
        updated_at = make_aware(updated_at) if updated_at else updated_at

        nominations = self.get_nominations(party, period, rows['EssenNom'])
        approvals = self.get_exemptions_approved(
            party, period, rows['EssenExemp']
        )
        approved_critical_uses = self.get_critical_uses_approved(
            rows['MeBrAgreedCriticalUseCategories']
        )

        if not created_at:
            created_at = updated_at
        elif not updated_at:
            updated_at = created_at

        return {
            "submission": {
                "schema_version": "legacy",
                "submitted_at": submitted_at,
                "created_at": created_at,
                "updated_at": updated_at,
                "version": 1,
                "_workflow_class": "default_exemption",
                "_current_state": "finalized",
                "_previous_state": None,
                "flag_provisional": False,
                "flag_valid": True,
                "flag_superseded": False,
                "created_by_id": self.admin.id,
                "last_edited_by_id": self.admin.id,
                "obligation_id": 11,
                "party_id": party.id,
                "reporting_period_id": period.id,
                "cloned_from_id": None,
                # "info_id": "",
                "reporting_channel": ReportingChannel.objects.get(name="Legacy")
            },
            "submission_info": {
                "reporting_officer": "",
                "designation": "",
                "organization": "",
                "postal_address": "",
                "country_id": party.id,
                "phone": "",
                "email": "",
                "date": created_at
            },
            "nominations": nominations,
            "exemptionapproveds": approvals,
            "approved_critical_uses": approved_critical_uses,
        }

    def get_rafs(self, party, period, rows):
        """
        Parses the RAF data for the submission identified by
        the party/period combination.

        Return a list of RAFs.
        """

        rafs = []

        for pk, entry in rows.items():
            # pk = ("SubstID", "EssenCrit")
            raf_entries = entry["raf"]
            critical_entries = entry["critical"]
            try:
                substance_id = self.substances[pk[0]].id
            except KeyError as e:
                logger.error("Reporting new unknown substance %s: %s/%s", e, party.abbr, period.name)
                continue

            rafs.append({
                "raf_report": {
                    "substance_id": substance_id,
                    # ImpSrcCntryID will be added into `imports` table
                    "quantity_exempted": sum(
                        float_to_decimal_zero_if_none(e["Exempted"])
                        for e in raf_entries
                    ),
                    "quantity_production": sum(
                        float_to_decimal_zero_if_none(e["Produced"])
                        for e in raf_entries
                    ),
                    "on_hand_start_year": sum(
                        float_to_decimal_zero_if_none(e["OpenBal"])
                        for e in raf_entries
                    ),
                    "quantity_used": sum(
                        float_to_decimal_zero_if_none(e["EssenUse"])
                        for e in raf_entries
                    ),
                    "quantity_exported": sum(
                        float_to_decimal_zero_if_none(e["Exported"])
                        for e in raf_entries
                    ),
                    "quantity_destroyed": sum(
                        float_to_decimal_zero_if_none(e["Destroyed"])
                        for e in raf_entries
                    ),
                    "is_emergency": pk[1],
                    "remarks_os": self.get_raf_remarks(raf_entries),
                    "remarks_party": ""
                },
                "imports": self.get_raf_imports(raf_entries),
                "use_categories": self.get_use_categories(critical_entries)
            })

        return rafs

    def get_raf_remarks(self, entries):
        remarks = []
        for entry in entries:
            try:
                party = self.parties[entry["ImpSrcCntryID"]].name
            except KeyError as e:
                party = None

            if entry["Remark"]:
                if party:
                    remarks.append(party + ": ")
                remarks.append(entry["Remark"])

        return '\n'.join(remarks)

    def get_raf_imports(self, entries):
        raf_imports = []
        for entry in entries:
            if entry["Imported"] is not None:
                logger.info(f'{entry["CntryID"]}/{entry["PeriodID"]}: processing import from {entry["ImpSrcCntryID"]}, quantity {entry["Imported"]}')
                try:
                    party = self.parties[entry["ImpSrcCntryID"]]
                except KeyError as e:
                    if entry["ImpSrcCntryID"] != 'UNK':
                        logger.error(
                            "RAF (imports): Unknown source party for %s/%s: %s",
                            entry["CntryID"],
                            entry["PeriodID"],
                            entry["ImpSrcCntryID"],
                        )
                    party = None
                raf_imports.append({
                    "party": party,
                    "quantity": float_to_decimal_zero_if_none(entry["Imported"])
                })
        return raf_imports

    def get_use_categories(self, entries):
        use_categories = []

        for entry in entries:
            logger.info(
                f'{entry["CntryID"]}/{entry["PeriodID"]}: processing critical '
                f'use category {entry["CU_Title"]}'
            )
            try:
                code = CriticalUseCategory.get_alt_name(entry['CU_Title'])
                category = CriticalUseCategory.objects.get(code=code)
            except CriticalUseCategory.DoesNotExist:
                logger.warning(
                    f'Unknown category {code} for {entry["CntryID"]}/'
                    f'{entry["PeriodID"]}'
                )
                category = None

            use_categories.append({
                'quantity': float_to_decimal(entry["CU_Amount"]),
                'critical_use_category': category,
            })

        return use_categories

    def check_is_emergency(self, rows):
        """
        If one entry is an emergency, then the whole submission becomes an emergency.
        """

        for row in rows:
            if row['IsEmergency']:
                return True

        return False

    def get_nominations(self, party, period, rows):
        """
        Parses the Exemption Nomination data for the submission identified by
        the party/period combination.

        Return a list of exemption nominations.
        """

        nominations = []

        for row in rows:
            try:
                substance_id = self.substances[row['SubstID']].id
            except KeyError as e:
                logger.error("Reporting new unknown substance %s: %s/%s", e, party.abbr, period.name)
                continue

            nominations.append({
                "substance_id": substance_id,
                "quantity": float_to_decimal(row['SubmitAmt']),
                "remarks_os": row['Remark'] if row['Remark'] else ""
            })

        return nominations

    def get_exemptions_approved(self, party, period, rows):
        """
        Parses the Exemption Approved data for the submission identified by
        the party/period combination.

        Return a list of approved exemptions.
        """

        approved_exemptions = []

        for row in rows:
            try:
                substance_id = self.substances[row['SubstID']].id
            except KeyError as e:
                logger.error("Reporting new unknown substance %s: %s/%s", e, party.abbr, period.name)
                continue

            approved_exemptions.append({
                "substance_id": substance_id,
                "decision_approved": row['ApprDec'],
                "approved_teap_amount": float_to_decimal(row['ApprTEAP']),
                "quantity": float_to_decimal(row["ApprAmt"]),
                "remarks_os": row['Remark'] if row['Remark'] else "",
                "is_emergency": True if row["IsEmergency"] else False
            })

        return approved_exemptions

    def get_critical_uses_approved(self, rows):
        """
        Returns a list of approved critical use categories from the rows.
        """

        approved_critical_uses = []

        for row in rows:

            approved_critical_uses.append({
                "decision": row['ApprDec'],
                "category_name": row['Categories of permitted critical uses'],
                "quantity": float_to_decimal(row["ApprAmt"]),
            })

        return approved_critical_uses

    def create_approved_critical_uses(self, party, period, items):
        for item in items:
            code = CriticalUseCategory.get_alt_name(item['category_name'])
            decision = item['decision']
            # Remove number in bracket when searching for the exemption
            # to handle inconsistent decision numbers
            decision_base = decision[:decision.find('(')]
            try:
                category = CriticalUseCategory.objects.get(code=code)
                exemption = ExemptionApproved.objects.get(
                    submission__party=party,
                    submission__reporting_period=period,
                    decision_approved__startswith=decision_base,
                    substance__has_critical_uses=True,
                    is_emergency=False,
                )
                ApprovedCriticalUse.objects.create(
                    exemption=exemption,
                    critical_use_category=category,
                    quantity=item['quantity']
                )
            except CriticalUseCategory.DoesNotExist:
                logger.error(
                    "Unknown category %s for %s/%s",
                    item['category_name'], party.abbr, period.name
                )
            except ExemptionApproved.DoesNotExist:
                logger.error(
                    "Unknown exemption: %s/%s/%s",
                    party.abbr, period.name, item['decision']
                )
