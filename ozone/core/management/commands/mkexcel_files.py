import warnings

import openpyxl
from django.core.management.base import BaseCommand

warnings.filterwarnings("ignore")

OUTPUT_EXCEL_FILES = {
    'initial_data': [
        'Regions',
        'SubRegion',
        'Cntry',
        'Subst',
        'Blends',
        'BlendComposition',
        'Period',
        'CntryYr',
        'MDG_RegionsCntryArea',
        'MeBrAgreedCriticalUseCategories',
        'MeBrActualCriticalUsebyCategory'
    ],
    'legacy_compliance': [
        'DeviationTypes',
        'DeviationSources',
        'Plans_of_Action',
        'Plans_of_Action_Decs'
    ],
    'legacy_exemptions': [
        'EssenUse',
        'EssenNom',
        'EssenExemp'
    ],
    'legacy_procagents': [
        'ProcAgentContanTechnology',
        'ProcAgentEmitLimits',
        'ProcAgentEmitLimitsValidity',
        'ProcAgentUses',
        'ProcAgentUsesReported',
        'ProcAgentUsesValidity',
        'ProcAgentUsesDateReported'
    ],
    'legacy_submissions': [
        'Overall',
        'Import',
        'ImportNew',
        'Export',
        'Produce',
        'Destroy',
        'NonPartyTrade',
        'NonPartyTradeNew'
    ],
    'legacy_transfers': [
        'ProdTransfersLetters',
        'Letters',
        'ProdTransfers'
    ],
    'tbl_prod_cons': [
        'tbl_ProdCons'
    ]
}


class Command(BaseCommand):
    help = 'Split the Excel master file by worksheets creating multiple Excel files'

    def add_arguments(self, parser):
        parser.add_argument('file', help="Master Excel file")

    def handle(self, *args, **options):
        source_workbook = openpyxl.load_workbook(filename=options['file'], read_only=True)
        self.stdout.write('Excel file loaded.')

        for name, sheets in OUTPUT_EXCEL_FILES.items():
            target_workbook = openpyxl.Workbook()
            target_workbook.remove_sheet(target_workbook.active)
            for sheet in sheets:
                source_worksheet = source_workbook.get_sheet_by_name(sheet)
                target_worksheet = target_workbook.create_sheet(sheet)

                for i_r, row in enumerate(source_worksheet.rows, start=1):
                    for i_c, cell in enumerate(row, start=1):
                        target_cell = target_worksheet.cell(column=i_c, row=i_r)
                        target_cell.value = cell.value
                        target_cell.number_format = cell.number_format or ''

            target_workbook.save('{name}.xlsx'.format(name=name))
            self.stdout.write(self.style.SUCCESS('Successfully generated "%s" Excel file' % name))
