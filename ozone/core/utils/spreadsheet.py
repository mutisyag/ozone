from collections import OrderedDict
import csv
import json
from datetime import datetime
import openpyxl


class DefaultConverter:

    def to_csv(self, value, column):
        return str(value)

    def from_csv(self, value, column):
        return value


class Table:

    converter_cls = DefaultConverter

    def __init__(self, header=[], rows=[]):
        self.header = list(header)
        self.rows = list(rows)

    @classmethod
    def from_xlsx_sheet(cls, sheet):
        values = list(sheet.values)
        headers = values[0]
        rows = (dict(zip(headers, row)) for row in values[1:])
        return cls(headers, rows)

    @classmethod
    def from_csv(cls, path):
        converter = cls.converter_cls()

        def format_row(row):
            rv = OrderedDict()
            for column, value in row.items():
                try:
                    rv[column] = converter.from_csv(value, column)
                except:
                    raise RuntimeError(f"{converter}.from_csv failed for "
                                       f"column={column} value={value!r}")
            return rv

        with path.open(encoding='utf8') as f:
            reader = csv.DictReader(f)
            rows = (format_row(r) for r in reader)
            return cls(reader.fieldnames, rows)

    def dump_xlsx_sheet(self, sheet):
        for (c, col) in enumerate(self.header, start=1):
            sheet.cell(row=1, column=c).value = col

        for (r, row) in enumerate(self.rows, start=2):
            for (c, col) in enumerate(self.header, start=1):
                sheet.cell(row=r, column=c).value = row[col]

    def dump_csv(self, path):
        converter = self.converter_cls()

        def format_row(row):
            rv = OrderedDict()
            for column, value in row.items():
                try:
                    rv[column] = converter.to_csv(value, column)
                except:
                    raise RuntimeError(f"{converter}.to_csv failed for "
                                       f"column={column} value={value!r}")
            return rv

        with path.open('w', encoding='utf8') as f:
            writer = csv.DictWriter(f, self.header)
            writer.writeheader()

            for row in self.rows:
                writer.writerow(format_row(row))


class Spreadsheet:

    table_cls = Table

    def __init__(self):
        self.tables = OrderedDict()

    @classmethod
    def from_xlsx(cls, path):
        wb = openpyxl.load_workbook(filename=path)
        spreadsheet = cls()

        for sheet in wb:
            table = spreadsheet.table_cls.from_xlsx_sheet(sheet)
            spreadsheet.tables[sheet.title] = table

        return spreadsheet

    @classmethod
    def from_csvdir(cls, path):
        with (path / '.csvdir.json').open(encoding='utf8') as f:
            meta = json.load(f)

        spreadsheet = cls()

        for name in meta['sheets']:
            table = spreadsheet.table_cls.from_csv(path / f"{name}.csv")
            spreadsheet.tables[name] = table

        return spreadsheet

    def dump_xlsx(self, path):
        workbook = openpyxl.Workbook()
        workbook.remove_sheet(workbook.active)

        for name, table in self.tables.items():
            table.dump_xlsx_sheet(workbook.create_sheet(name))

        workbook.save(path)

    def dump_csvdir(self, path):
        path.mkdir(parents=True, exist_ok=True)

        for name, table in self.tables.items():
            table.dump_csv(path / f"{name}.csv")

        meta = {'sheets': list(self.tables)}
        with (path / '.csvdir.json').open('w', encoding='utf8') as f:
            json.dump(meta, f, sort_keys=True, indent=2)
            f.write("\n")


class OzoneConverter(DefaultConverter):

    boolean = [
        'AI_ComplRep',
        'AII_ComplRep',
        'BI_ComplRep',
        'BII_ComplRep',
        'BIII_ComplRep',
        'CI_ComplRep',
        'CII_ComplRep',
        'CIII_ComplRep',
        'EI_ComplRep',
        'F_ComplRep',
        'Checked_Blanks',
    ]

    numeric = [
        'DataID',
        'Imported',
        'Exported',
        'Produced',
        'Destroyed',
        'NonPartyTrade',
        'Emitted',
        'SubstID',
        'ImpNew',
        'ImpRecov',
        'ImpFeedstock',
        'ImpEssenUse',
        'ImpLabUse',
        'ImpQuarAppl',
        'ImpProcAgent',
        'ImpPolyol',
        'NPTImpNew',
        'NPTImpRecov',
        'NPTExpNew',
        'ProdAllNew',
        'ProdFeedstock',
        'ProdEssenUse',
        'ProdLabUse',
        'ProdArt5',
        'ProdQuarAppl',
        'ProdProcAgent',
    ]

    date = [
        'DateReported',
        'DateCreate',
        'DateUpdate',
    ]

    def from_csv(self, value, column):
        if column in self.boolean:
            if value == '':
                return None

            return int(value)

        if column in self.numeric:
            if value == '':
                return None

            return float(value)

        if column in self.date:
            if value == '':
                return None

            return datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%f')

        return value

    def to_csv(self, value, column):
        if value is None:
            return ''

        if column in self.date:
            return value.strftime('%Y-%m-%dT%H:%M:%S.%f')

        return value


class OzoneTable(Table):

    converter_cls = OzoneConverter


class OzoneSpreadsheet(Spreadsheet):

    table_cls = OzoneTable
